import openpyxl
import glob
import os
from openpyxl.styles import Color, PatternFill, Font, Border, Side

octant_sign = [1, -1, 2, -2, 3, -3, 4, -4]
octant_name_id_mapping = {1: "Internal outward interaction", -1: "External outward interaction", 2: "External Ejection",
                          -2: "Internal Ejection", 3: "External inward interaction", -3: "Internal inward interaction",
                          4: "Internal sweep", -4: "External sweep"}
yellow_color = "00FFFF00"
bg_yellow = PatternFill(start_color=yellow_color, end_color=yellow_color, fill_type='solid')
black_color = "00000000"
double = Side(border_style="thin", color=black_color)
black_border = Border(top=double, left=double, right=double, bottom=double)

def reset_count(count_seq):
    for x in octant_sign:
        count_seq[x] = 0

def reset_count_except(count_seq, left):
    for x in octant_sign:
        if (x != left):
            count_seq[x] = 0

def set_frequency(longest, frequency, outputSheet):
    
    for i in range(9):
        for j in range(3):
            outputSheet.cell(row=3 + i, column=45 + j).border = black_border

    outputSheet.cell(row=3, column=45).value = "Octant ##"
    outputSheet.cell(row=3, column=46).value = "Longest Subsquence Length"
    outputSheet.cell(row=3, column=47).value = "count_seq"

    for i, label in enumerate(octant_sign):
        current_row = i + 3
        outputSheet.cell(row=current_row + 1, column=45).value = label
        outputSheet.cell(column=46, row=current_row + 1).value = longest[label]
        outputSheet.cell(column=47, row=current_row + 1).value = frequency[label]

def longest_subsequence_time(longest, frequency, timeRange, outputSheet):
    
    col_length = 50
    col_freq = 51

    
    row = 4

    outputSheet.cell(row=3, column=49).value = "Octant ###"
    outputSheet.cell(row=3, column=50).value = "Longest Subsquence Length"
    outputSheet.cell(row=3, column=51).value = "count_seq"

    
    for octant in octant_sign:
        
        outputSheet.cell(column=49, row=row).value = octant
        outputSheet.cell(column=col_length, row=row).value = longest[octant]
        outputSheet.cell(column=col_freq, row=row).value = frequency[octant]
    

        row += 1

        outputSheet.cell(column=49, row=row).value = "Time"
        outputSheet.cell(column=col_length, row=row).value = "From"
        outputSheet.cell(column=col_freq, row=row).value = "To"


        row += 1

        
        for timeData in timeRange[octant]:
            
            outputSheet.cell(row=row, column=col_length).value = timeData[0]
            outputSheet.cell(row=row, column=col_freq).value = timeData[1]

            row += 1

    for i in range(3, row):
        for j in range(49, 52):
            outputSheet.cell(row=i, column=j).border = black_border

def count_longest_subsequence_freq_func(longest, outputSheet, total_count):
    
    count_seq = {}

    
    frequency = {}

    
    timeRange = {}

    for label in octant_sign:
        timeRange[label] = []

    
    reset_count(count_seq)
    reset_count(frequency)

    
    last = -10

    
    for i in range(0, total_count):
        current_row = i + 3

        current = int(outputSheet.cell(column=11, row=current_row).value)

       
        if (current == last):
            count_seq[current] += 1
        else:
            count_seq[current] = 1
            reset_count_except(count_seq, current)

        
        if (count_seq[current] == longest[current]):
            frequency[current] += 1

            
            end = float(outputSheet.cell(row=current_row, column=1).value)
            start = 100 * end - longest[current] + 1
            start /= 100

            
            timeRange[current].append([start, end])

            
            reset_count(count_seq)
        else:
            reset_count_except(count_seq, current)

        
        last = current

    
    set_frequency(longest, frequency, outputSheet)

    
    longest_subsequence_time(longest, frequency, timeRange, outputSheet)

    print(longest)
    print(frequency)

def find_longest_subsequence(outputSheet, total_count):
    
    count_seq = {}

    longest = {}

    
    reset_count(count_seq)
    reset_count(longest)

    
    last = -10

   
    for i in range(0, total_count):
        current_row = i + 3

        current = int(outputSheet.cell(column=11, row=current_row).value)

       
        if (current == last):
            count_seq[current] += 1
            longest[current] = max(longest[current], count_seq[current])
            reset_count_except(count_seq, current)
        else:
            count_seq[current] = 1
            longest[current] = max(longest[current], count_seq[current])
            reset_count_except(count_seq, current)

        
        last = current

    
    count_longest_subsequence_freq_func(longest, outputSheet, total_count)

def transition_count_func(row, transition_count, outputSheet):
   

    outputSheet.cell(row=row, column=36).value = "To"
    outputSheet.cell(row=row + 1, column=35).value = "Octant #"
    outputSheet.cell(row=row + 2, column=34).value = "From"

    for i in range(35, 44):
        for j in range(row + 1, row + 1 + 9):
            outputSheet.cell(row=j, column=i).border = black_border


    
    for i, label in enumerate(octant_sign):

        outputSheet.cell(row=row + 1, column=i + 36).value = label
        outputSheet.cell(row=row + i + 2, column=35).value = label


    
    for i, l1 in enumerate(octant_sign):
        maxi = -1

        for j, l2 in enumerate(octant_sign):
            val = transition_count[str(l1) + str(l2)]
            maxi = max(maxi, val)

        for j, l2 in enumerate(octant_sign):

            outputSheet.cell(row=row + i + 2, column=36 + j).value = transition_count[str(l1) + str(l2)]
            if transition_count[str(l1) + str(l2)] == maxi:
                maxi = -1
                outputSheet.cell(row=row + i + 2, column=36 + j).fill = bg_yellow

def set_mod_overall_transition_count(outputSheet, mod, total_count):
    

    total_partition = total_count // mod

    if (total_count % mod != 0):
        total_partition += 1

    
    rowStart = 16

   
    for i in range(0, total_partition):
       
        start = i * mod
        end = min((i + 1) * mod - 1, total_count - 1)

        

        outputSheet.cell(column=35, row=rowStart - 1 + 13 * i).value = "Mod Transition count_seq"
        outputSheet.cell(column=35, row=rowStart + 13 * i).value = str(start) + "-" + str(end)


        
        transition_count = {}
        for a in range(1, 5):
            for b in range(1, 5):
                transition_count[str(a) + str(b)] = 0
                transition_count[str(a) + str(-b)] = 0
                transition_count[str(-a) + str(b)] = 0
                transition_count[str(-a) + str(-b)] = 0

        
        for a in range(start, end + 1):

            current = outputSheet.cell(column=11, row=a + 3).value
            next = outputSheet.cell(column=11, row=a + 4).value


            if (next != None):
                transition_count[str(current) + str(next)] += 1

       
        transition_count_func(rowStart + 13 * i, transition_count, outputSheet)
        print(transition_count)
def set_overall_Transition_Count(outputSheet, total_count):
    
    count_transition = {}
    for i in range(1, 5):
        for j in range(1, 5):
            count_transition[str(i) + str(j)] = 0
            count_transition[str(i) + str(-j)] = 0
            count_transition[str(-i) + str(j)] = 0
            count_transition[str(-i) + str(-j)] = 0

    
    start = 0

    

    last = int(outputSheet["K3"].value)


    while (start < total_count - 1):
        

        current = int(outputSheet.cell(row=start + 4, column=11).value)
        count_transition[str(last) + str(current)] += 1
        last = current

        start = start + 1

   
    transition_count_func(2, count_transition, outputSheet)
    print(count_transition)

def set_rank_count(row, countMap, outputSheet):
    
    sorted_count = []
    count_seq = []
    for label in octant_sign:
        count_seq.append(countMap[label])

    for ct in count_seq:
        sorted_count.append(ct)

    sorted_count.sort(reverse=True)

    rank = []

    for i, el in enumerate(count_seq):
        for j, ell in enumerate(sorted_count):
            if (ell == el):
                rank.append(j + 1)
                sorted_count[j] = -1
                break
    first_rank_oct = -10

    for j in range(0, 8):
        outputSheet.cell(row=row, column=23 + j).value = rank[j]
        if (rank[j] == 1):
            first_rank_oct = octant_sign[j]
            outputSheet.cell(row=row, column=23 + j).fill = bg_yellow

    outputSheet.cell(row=row, column=31).value = first_rank_oct
    outputSheet.cell(row=row, column=32).value = octant_name_id_mapping[first_rank_oct]

def overall_octant_rank_func(last_row, outputSheet):
    count_seq = {-1: 0, 1: 0, -2: 0, 2: 0, -3: 0, 3: 0, -4: 0, 4: 0}

    row = 4
    while outputSheet.cell(row=row, column=29).value is not None:
        octant = int(outputSheet.cell(row=row, column=31).value)
        count_seq[octant] += 1
        row += 1

    for i in range(9):
        for j in range(3):
            row = last_row + 2 + i
            col = 29 + j
            outputSheet.cell(row=row, column=col).border = black_border

    outputSheet.cell(column=29, row=last_row + 2).value = "Octant ID"
    outputSheet.cell(column=30, row=last_row + 2).value = "Octant Name "
    outputSheet.cell(column=31, row=last_row + 2).value = "count_seq of Rank 1 Mod Values"

    for j, octant in enumerate(octant_sign):
        outputSheet.cell(column=29, row=last_row + 3 + j).value = octant
        outputSheet.cell(column=30, row=last_row + 3 + j).value = octant_name_id_mapping[octant]
        outputSheet.cell(column=31, row=last_row + 3 + j).value = count_seq[octant]

    print(count_seq)

def set_mod_count(outputSheet, mod, total_count):
    
    count_seq = {-1: 0, 1: 0, -2: 0, 2: 0, -3: 0, 3: 0, -4: 0, 4: 0}

   
    last_row = -1

    
    start = 0
    while (start < total_count):

        count_seq[int(outputSheet.cell(row=start + 3, column=11).value)] += 1


        start += 1

        if (start % mod == 0):
           

            row = 4 + start // mod
            last_row = row
            outputSheet.cell(row=row, column=14).value = str(start - mod) + "-" + str(
                min(total_count, start - 1))

            for i, label in enumerate(octant_sign):
                outputSheet.cell(row=row, column=15 + i).value = count_seq[label]

            set_rank_count(row, count_seq, outputSheet)
            print(count_seq)


            
            count_seq = {-1: 0, 1: 0, -2: 0, 2: 0, -3: 0, 3: 0, -4: 0, 4: 0}


    if (start % mod != 0):
        

        row = 5 + start // mod
        last_row = row
        outputSheet.cell(row=row, column=14).value = str(start - mod) + "-" + str(min(total_count, start - 1))
        for i, label in enumerate(octant_sign):
            outputSheet.cell(row=row, column=15 + i).value = count_seq[label]

        set_rank_count(row, count_seq, outputSheet)


    if (last_row != -1):
        overall_octant_rank_func(last_row, outputSheet)

def setOverallCount(total_count, outputSheet):
   
    count_seq = {-1: 0, 1: 0, -2: 0, 2: 0, -3: 0, 3: 0, -4: 0, 4: 0}
    

    for i in range(3, total_count + 3):
        count_seq[int(outputSheet.cell(column=11, row=i).value)] = count_seq[int(outputSheet.cell(column=11,
                                                                                            row=i).value)] + 1


    
    for i, label in enumerate(octant_sign):

        outputSheet.cell(row=4, column=i + 15).value = count_seq[label]


    set_rank_count(4, count_seq, outputSheet)
    print(count_seq)


def set_overall_octant_rank_count(outputSheet, mod, total_count):
    headers = ["Octant ID", 1, -1, 2, -2, 3, -3, +4, -4, "Rank Octant 1", "Rank Octant -1", "Rank Octant 2",
               "Rank Octant -2", "Rank Octant 3", "Rank Octant -3", "Rank Octant 4", "Rank Octant -4",
               "Rank1 Octant ID", "Rank1 Octant Name"]

    total_rows = total_count // mod + 1 + 1  
    if total_count % mod != 0:
        total_rows += 1

    for i, header in enumerate(headers):
        for j in range(total_rows):
            outputSheet.cell(row=3 + j, column=14 + i).border = black_border

    for i, header in enumerate(headers):
        outputSheet.cell(row=3, column=i + 14).value = header

    outputSheet.cell(row=4, column=13).value = "Mod " + str(mod)

    setOverallCount(total_count, outputSheet)

def get_octant(x, y, z):
    if (x >= 0 and y >= 0):
        if (z >= 0):
            return 1
        else:
            return -1

    if (x < 0 and y >= 0):
        if (z >= 0):
            return 2
        else:
            return -2

    if (x < 0 and y < 0):
        if (z >= 0):
            return 3
        else:
            return -3

    if (x >= 0 and y < 0):
        if (z >= 0):
            return 4
        else:
            return -4

def setProcessedDataWithOctant(u_avg, v_avg, w_avg, total_count, inputSheet, outputSheet):
    start = 2
    time = inputSheet.cell(start, 1).value

    
    while (time != None):
        

        u1 = inputSheet.cell(start, 2).value - u_avg
        v1 = inputSheet.cell(start, 3).value - v_avg
        w1 = inputSheet.cell(start, 4).value - w_avg

        u1 = round(u1, 3)
        v1 = round(v1, 3)
        w1 = round(w1, 3)

        octant = get_octant(u1, v1, w1)


       

        outputSheet.cell(row=start + 1, column=8).value = u1
        outputSheet.cell(row=start + 1, column=9).value = v1
        outputSheet.cell(row=start + 1, column=10).value = w1
        outputSheet.cell(row=start + 1, column=11).value = octant


        start = start + 1

        time = inputSheet.cell(start, 1).value
        print(u1, v1, w1)

def set_input_data(input_file_name, outputSheet):
    wb = openpyxl.load_workbook(input_file_name)
    inputSheet = wb.active

    start = 2
    time = inputSheet.cell(start, 1).value

    
    u_sum = 0
    v_sum = 0
    w_sum = 0

    
    while (time != None):

        u_sum += float(inputSheet.cell(start, 2).value)
        v_sum += float(inputSheet.cell(start, 3).value)
        w_sum += float(inputSheet.cell(start, 4).value)


        
        outputSheet.cell(row=start + 1, column=1).value = inputSheet.cell(start, 1).value
        outputSheet.cell(row=start + 1, column=2).value = inputSheet.cell(start, 2).value
        outputSheet.cell(row=start + 1, column=3).value = inputSheet.cell(start, 3).value
        outputSheet.cell(row=start + 1, column=4).value = inputSheet.cell(start, 4).value


        start = start + 1
        time = inputSheet.cell(start, 1).value

    
    total_count = start - 2  # -1 for header and -1 for last None
   

    u_avg = round(u_sum / total_count, 3)
    v_avg = round(v_sum / total_count, 3)
    w_avg = round(w_sum / total_count, 3)


    

    outputSheet.cell(row=3, column=5).value = u_avg
    outputSheet.cell(row=3, column=6).value = v_avg
    outputSheet.cell(row=3, column=7).value = w_avg


    
    setProcessedDataWithOctant(u_avg, v_avg, w_avg, total_count, inputSheet, outputSheet)
    print(u_avg, v_avg, w_avg)

    return total_count

def entry_point(wb, mod):
    fileName = wb.split("/")[-1]
    fileName = fileName.split(".xlsx")[0]
    outputFileName = "output/" + fileName + "_octant_analysis_mod_" + str(mod) + ".xlsx"

    outputFile = openpyxl.Workbook()
    outputSheet = outputFile.active

    outputSheet.cell(row=1, column=14).value = "Overall Octant count_seq"
    outputSheet.cell(row=1, column=24).value = "Rank #1 Should be highlighted Yellow"
    outputSheet.cell(row=1, column=35).value = "Overall Transition count_seq"
    outputSheet.cell(row=1, column=45).value = "Longest Subsequence Length"
    outputSheet.cell(row=1, column=49).value = "Longest Subsequence Length with Range"
    outputSheet.cell(row=2, column=36).value = "To"

    headers = ["T", "U", "V", "W", "U Avg", "V Avg", "W Avg", "U'=U - U avg", "V'=V - V avg", "W'=W - W avg", "Octant"]
    for i, header in enumerate(headers):
        outputSheet.cell(row=2, column=i + 1).value = header

    total_count = set_input_data(wb, outputSheet)
    set_overall_octant_rank_count(outputSheet, mod, total_count)
    set_mod_count(outputSheet, mod, total_count)
    set_overall_Transition_Count(outputSheet, total_count)
    set_mod_overall_transition_count(outputSheet, mod, total_count)
    find_longest_subsequence(outputSheet, total_count)

    outputFile.save(outputFileName)

def octant_analysis(mod=5000):
    path = os.getcwd()
    # csv_files = glob.glob(os.path.join(path + "\input", "*.xlsx"))
    csv_files = glob.glob(os.path.join("input/","*.xlsx"))

    for file in csv_files:
        entry_point(file, mod)

mod = 5000
octant_analysis(mod)