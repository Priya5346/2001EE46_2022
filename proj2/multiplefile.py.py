
import streamlit as st
from datetime import datetime
import os
import pandas as pd
import openpyxl as op
import glob
from openpyxl.styles import Border,Side
os.chdir(r'C:\Users\PRIYA RAJ\OneDrive\Documents\GitHub\2001EE46_2022\proj2')

c=0
temp=0
cols=['+1','-1','+2','-2','+3','-3','+4','-4']

def transcount(start,end,df):
    """function to count the transition of octants for a given range of a given file - It returns a list of lists containing the transition count for each octant

    Args:
        start (int): start of the range
        end (int): end of the range
        df (file): file to operate on
    """
    try:
        if end==c:                                      #checking for the end condition to eliminate the last case transition problem
            end-=1
        table=[[0]*9 for _ in range(8)]                 #creating the table for the transition count 
        for i in range(0,8):                            #assigning first column of the table with octants
            table[i][0]=cols[i]
        for i in range(start,end):                      #starting the loop for the given range

            if df.loc[i,'Octant']==1:                   #now checking for transitions of each octant one by one
                if df.loc[i+1,'Octant']==1:
                    table[0][1]+=1
                elif df.loc[i+1,'Octant']==-1:
                    table[0][2]+=1
                elif df.loc[i+1,'Octant']==2:
                    table[0][3]+=1
                elif df.loc[i+1,'Octant']==-2:
                    table[0][4]+=1
                elif df.loc[i+1,'Octant']==3:
                    table[0][5]+=1
                elif df.loc[i+1,'Octant']==-3:
                    table[0][6]+=1
                elif df.loc[i+1,'Octant']==4:
                    table[0][7]+=1
                elif df.loc[i+1,'Octant']==-4:
                    table[0][8]+=1

            elif df.loc[i,'Octant']==-1:
                if df.loc[i+1,'Octant']==1:
                    table[1][1]+=1
                elif df.loc[i+1,'Octant']==-1:
                    table[1][2]+=1
                elif df.loc[i+1,'Octant']==2:
                    table[1][3]+=1
                elif df.loc[i+1,'Octant']==-2:
                    table[1][4]+=1
                elif df.loc[i+1,'Octant']==3:
                    table[1][5]+=1
                elif df.loc[i+1,'Octant']==-3:
                    table[1][6]+=1
                elif df.loc[i+1,'Octant']==4:
                    table[1][7]+=1
                elif df.loc[i+1,'Octant']==-4:
                    table[1][8]+=1

            elif df.loc[i,'Octant']==2:
                if df.loc[i+1,'Octant']==1:
                    table[2][1]+=1
                elif df.loc[i+1,'Octant']==-1:
                    table[2][2]+=1
                elif df.loc[i+1,'Octant']==2:
                    table[2][3]+=1
                elif df.loc[i+1,'Octant']==-2:
                    table[2][4]+=1
                elif df.loc[i+1,'Octant']==3:
                    table[2][5]+=1
                elif df.loc[i+1,'Octant']==-3:
                    table[2][6]+=1
                elif df.loc[i+1,'Octant']==4:
                    table[2][7]+=1
                elif df.loc[i+1,'Octant']==-4:
                    table[2][8]+=1

            elif df.loc[i,'Octant']==-2:
                if df.loc[i+1,'Octant']==1:
                    table[3][1]+=1
                elif df.loc[i+1,'Octant']==-1:
                    table[3][2]+=1
                elif df.loc[i+1,'Octant']==2:
                    table[3][3]+=1
                elif df.loc[i+1,'Octant']==-2:
                    table[3][4]+=1
                elif df.loc[i+1,'Octant']==3:
                    table[3][5]+=1
                elif df.loc[i+1,'Octant']==-3:
                    table[3][6]+=1
                elif df.loc[i+1,'Octant']==4:
                    table[3][7]+=1
                elif df.loc[i+1,'Octant']==-4:
                    table[3][8]+=1

            elif df.loc[i,'Octant']==3:
                if df.loc[i+1,'Octant']==1:
                    table[4][1]+=1
                elif df.loc[i+1,'Octant']==-1:
                    table[4][2]+=1
                elif df.loc[i+1,'Octant']==2:
                    table[4][3]+=1
                elif df.loc[i+1,'Octant']==-2:
                    table[4][4]+=1
                elif df.loc[i+1,'Octant']==3:
                    table[4][5]+=1
                elif df.loc[i+1,'Octant']==-3:
                    table[4][6]+=1
                elif df.loc[i+1,'Octant']==4:
                    table[4][7]+=1
                elif df.loc[i+1,'Octant']==-4:
                    table[4][8]+=1

            elif df.loc[i,'Octant']==-3:
                if df.loc[i+1,'Octant']==1:
                    table[5][1]+=1
                elif df.loc[i+1,'Octant']==-1:
                    table[5][2]+=1
                elif df.loc[i+1,'Octant']==2:
                    table[5][3]+=1
                elif df.loc[i+1,'Octant']==-2:
                    table[5][4]+=1
                elif df.loc[i+1,'Octant']==3:
                    table[5][5]+=1
                elif df.loc[i+1,'Octant']==-3:
                    table[5][6]+=1
                elif df.loc[i+1,'Octant']==4:
                    table[5][7]+=1
                elif df.loc[i+1,'Octant']==-4:
                    table[5][8]+=1

            elif df.loc[i,'Octant']==4:
                if df.loc[i+1,'Octant']==1:
                    table[6][1]+=1
                elif df.loc[i+1,'Octant']==-1:
                    table[6][2]+=1
                elif df.loc[i+1,'Octant']==2:
                    table[6][3]+=1
                elif df.loc[i+1,'Octant']==-2:
                    table[6][4]+=1
                elif df.loc[i+1,'Octant']==3:
                    table[6][5]+=1
                elif df.loc[i+1,'Octant']==-3:
                    table[6][6]+=1
                elif df.loc[i+1,'Octant']==4:
                    table[6][7]+=1
                elif df.loc[i+1,'Octant']==-4:
                    table[6][8]+=1

            elif df.loc[i,'Octant']==-4:
                if df.loc[i+1,'Octant']==1:
                    table[7][1]+=1
                elif df.loc[i+1,'Octant']==-1:
                    table[7][2]+=1
                elif df.loc[i+1,'Octant']==2:
                    table[7][3]+=1
                elif df.loc[i+1,'Octant']==-2:
                    table[7][4]+=1
                elif df.loc[i+1,'Octant']==3:
                    table[7][5]+=1
                elif df.loc[i+1,'Octant']==-3:
                    table[7][6]+=1
                elif df.loc[i+1,'Octant']==4:
                    table[7][7]+=1
                elif df.loc[i+1,'Octant']==-4:
                    table[7][8]+=1

        return table                                        #returning the table containing the transition counts
    except:
        print("Error in transition count function")         
        exit()


def octcount(start,end,df):
    """returns the list containing the count of each octant for the given range of a given file
    Args:
        start (int): start of the range
        end (int): end of the range
        df (file) :  file to operate on
    """
    try:
        octant_name_id_mapping = {1:"Internal outward interaction", -1:"External outward interaction", 2:"External Ejection", -2:"Internal Ejection", 3:"External inward interaction", -3:"Internal inward interaction", 4:"Internal sweep", -4:"External sweep"}
        octcnt=[[0]*8,[0]*10]          #first list for octant count and second list for rank of octants having last 2 elements as rank1 octant ID and its name
        for i in range(start,end):
            if df.loc[i,'Octant']==1:
                octcnt[0][0]+=1
            elif df.loc[i,'Octant']==-1:
                octcnt[0][1]+=1
            elif df.loc[i,'Octant']==2:
                octcnt[0][2]+=1
            elif df.loc[i,'Octant']==-2:
                octcnt[0][3]+=1
            elif df.loc[i,'Octant']==3:
                octcnt[0][4]+=1
            elif df.loc[i,'Octant']==-3:
                octcnt[0][5]+=1
            elif df.loc[i,'Octant']==4:
                octcnt[0][6]+=1
            elif df.loc[i,'Octant']==-4:
                octcnt[0][7]+=1
        #calculating ranks of the octants alongwith rank1 octant ID and its name
        temp=octcnt[0].copy()                       
        index={0:1,1:-1,2:2,3:-2,4:3,5:-3,6:4,7:-4}
        temp.sort(reverse=True)
        for i in range(8):
            for j in range(8):
                try:
                    if(temp[i]==octcnt[0][j]):
                        if i==0:
                            octcnt[1][8]=index[j]
                            octcnt[1][9]=octant_name_id_mapping[index[j]]
                        octcnt[1][j]=i+1
                except:
                    print("error while calculating ranks")
                    exit()
    except:
        print("error in octcount function")
        exit()
    return octcnt


def octant_analysis(path,mod=5000):
    try:
        print("code started, wait for all the files to be processed")
        myfiles = glob.glob(os.path.join(path,"*.xlsx"))
        st.success("Please wait for the output message")
        for file in myfiles:
            print("starting to work on a new file")
            df=pd.read_excel(file)
            outputpath=path+"\\output"
            # file name
            now=datetime.now()
            fname = os.path.basename(file)
            file_name  = ""
            try:
                i=0
                sz=len(fname)
                while i<(sz-5):
                    file_name+=fname[i]
                    i+=1
            except:
                print("Program is strictly written for .xlsx extension files ")
            file_name+="_"+str(mod)+"_"+str(now.year)+"_"+str(now.month)+"_"+str(now.day)+"_"+str(now.hour)+"_"+str(now.minute)+"_"+str(now.second)+".xlsx"
            #reading the excel file
            meanu=df['U'].mean().round(decimals=3)    #mean of column U
            meanv=df['V'].mean().round(decimals=3)    #mean of column V
            meanw=df['W'].mean().round(decimals=3)    #mean of column W
            df['U Avg']=meanu
            df.loc[1:,'U Avg']=''
            df['V Avg']=meanv
            df.loc[1:,'V Avg']=''
            df['W Avg']=meanw
            df.loc[1:,'W Avg']=''
            df['''U'= U-U Avg''']=df['U'].round(decimals=3)-meanu         #making column of U'
            df['''V'= V-V Avg''']=df['V'].round(decimals=3)-meanv         #making column of V'
            df['''W'= W-W Avg''']=df['W'].round(decimals=3)-meanw         #making column of W'

            df["Octant"]=""                 #creating column 'octant'
            global c
            c=len(df.index)                                   #calculating number of rows
            modcount=int(c/mod)                         #calculating mod divisions
            cnt=[]                                      #list of octant counts for each mod
            for i in range(c):                          #assigning octant value
                x=df.loc[i,'''U'= U-U Avg''']
                y=df.loc[i,'''V'= V-V Avg''']
                z=df.loc[i,'''W'= W-W Avg''']
                if x >= 0 and y >= 0 and z >= 0:
                    df.loc[i,'Octant']=1
                elif x >= 0 and y >= 0 and z < 0:
                    df.loc[i,'Octant']=-1     
                elif x <= 0 and y >= 0 and z >= 0:
                    df.loc[i,'Octant']=2
                elif x <= 0 and y >= 0 and z < 0:
                    df.loc[i,'Octant']=-2
                elif x <= 0 and y < 0 and z >= 0:
                    df.loc[i,'Octant']=3
                elif x <= 0 and y < 0 and z < 0:
                    df.loc[i,'Octant']=-3
                elif x >= 0 and y < 0 and z >= 0:
                    df.loc[i,'Octant']=4
                elif x >= 0 and y < 0 and z < 0:
                    df.loc[i,'Octant']=-4
            df['']=''                                   #creating blank column
            df[' ']=''
            df.loc[1,' ']='Mod {}'.format(mod)           #inserting the 'mod {}' cell
            df[['Octant ID',1,-1,2,-2,3,-3,4,-4]]=''    #creating remaining columns
            df[['rank of 1','rank of -1','rank of 2','rank of -2','rank of 3','rank of -3','rank of 4','rank of -4','rank 1 Octant ID','rank 1 Octant Name']]=''
            allcnt=octcount(0,c,df)                     #calculating overall octant counts
            df.loc[0,'Octant ID']='Overall Count'   
            df.loc[0,1:-4]=allcnt[0]                    #writing the overall octant counts
            df.loc[0,'rank of 1':]=allcnt[1]            #writing the overall octant ranks    
            rnk1octid=[]                                #list of rank 1 Octant ID column
            cnt1=[]
            for i in range(0,modcount+1):               #calculating octant count and octant rank for each mod division and storing in list cnt[] and cnt1[] respectively
                if (i!=modcount):                       #octant count and ranks for all divisions except last
                    cnt.append(octcount(mod*i,mod*(i+1),df)[0])
                    cnt1.append(octcount(mod*i,mod*(i+1),df)[1])
                elif (i==modcount):                     #octant counts and ranks for last division
                    cnt.append(octcount(mod*i,c,df)[0])
                    cnt1.append(octcount(mod*i,c,df)[1])
            for i in range(0,modcount+1):               #creating and writing the table of octant counts and ranks for mod divisions
                if (i!=modcount):
                    df.loc[(1+i),'Octant ID']="{} - {}".format(mod*i,(mod*(i+1))-1)
                    cntmod=cnt[i].copy()                #copying the octant count for each division to a new list
                    cntmod1=cnt1[i].copy()              #copying the octant rank for each division to a new list
                    rnk1octid.append(cntmod1[8])
                    for j in range(14,22):              #assigning the octatnt count values to the cell
                        df.iloc[1+i,j]=cntmod[j-14]
                    for j in range(22,32):              #assigning the octatnt rank values to the cell
                        df.iloc[1+i,j]=cntmod1[j-22]
                elif(i==modcount):                      #repeating the process for last division
                    df.loc[(1+i),'Octant ID']="{} - {}".format(mod*i,c-1)
                    cntmod=cnt[i].copy()                #copying the octant count for each division to a new list
                    cntmod1=cnt1[i].copy()              #copying the octant rank for each division to a new list
                    rnk1octid.append(cntmod1[8])
                    for j in range(14,22):              #assigning the octant count values to the cell
                        df.iloc[1+i,j]=cntmod[j-14]
                    for j in range(22,32):              #assigning the octant rank values to the cell
                        df.iloc[1+i,j]=cntmod1[j-22]
            #calculating and writing the Count of Rank 1 Mod Values table
            index={0:1,1:-1,2:2,3:-2,4:3,5:-3,6:4,7:-4}
            octant_name_id_mapping = {1:"Internal outward interaction", -1:"External outward interaction", 2:"External Ejection", -2:"Internal Ejection", 3:"External inward interaction", -3:"Internal inward interaction", 4:"Internal sweep", -4:"External sweep"}
            cntrnk1mod=[0]*8                            #list for storing Count of Rank 1 Mod Values
            for i in range(8):
                for j in range(modcount+1):
                    if index[i]==rnk1octid[j]:
                        cntrnk1mod[i]+=1
            df.loc[6+modcount,'rank of 4':'rank 1 Octant ID']=['Octant ID','Octant Name','Count of Rank 1 Mod Values']
            for i in range(8):
                df.loc[7+modcount+i,'rank of 4':'rank 1 Octant ID']=[index[i],octant_name_id_mapping[index[i]],cntrnk1mod[i]]
            try:
                r=0
                df['  ']=''    #2 space blank column after ranking table
                df['   ']=''    #3 space blank column containing 'from' keyword
                df['Overall Transition Count']=''
                df['    ']=''            #4 space column
                df['     ']=''           #5 space column
                df['      ']=''          #6 space column
                df['       ']=''         #7 space column
                df['        ']=''        #8 space column
                df['         ']=''       #9 space column
                df['          ']=''      #10 space column
                df['           ']=''     #11 space column
                df.loc[r,'    ']='To'
                df.loc[r+2,'   ']='From'
                df.loc[r+1,'Overall Transition Count':]=['Octant#','+1','-1','+2','-2','+3','-3','+4','-4']
                table=transcount(0,c,df)
                for i in range(8):
                    df.loc[r+2+i,'Overall Transition Count':]=table[i]
                r+=(2+i)
            except:
                print("error in writing overall transition count table to file")
                exit()
            try:
                for i in range(0,modcount+1):
                    if(i!=modcount):
                        r+=3                            #increasing the row counter to create gap for the next table
                        df.loc[r,'Overall Transition Count']='Mod Transition Count'
                        df.loc[r+1,'Overall Transition Count']='{}-{}'.format(mod*i,(mod*(i+1))-1)
                        df.loc[r+1,'    ']='To'
                        df.loc[r+3,'   ']='From'
                        df.loc[r+2,'Overall Transition Count':]=['Octant#','+1','-1','+2','-2','+3','-3','+4','-4']         #creating the header of the table
                        table=transcount(mod*i,mod*(i+1),df)         #calculating the transitions for (i+1)th division using transcount function
                        for i in range (0,8):
                            df.loc[r+3+i,'Overall Transition Count':]=table[i]      #filling the created table with calculated values
                        r+=(3+i)                        #increasing row counter for the next table
                    elif(i==modcount):                  #repeating above process for the last division
                        r+=3
                        df.loc[r,'Overall Transition Count']='Mod Transition Count'
                        df.loc[r+1,'Overall Transition Count']='{}-{}'.format(mod*i,c-1)
                        df.loc[r+1,'    ']='To'
                        df.loc[r+3,'   ']='From'
                        df.loc[r+2,'Overall Transition Count':]=['Octant#','+1','-1','+2','-2','+3','-3','+4','-4']
                        table=transcount(mod*i,c,df)
                        for i in range (0,8):
                            df.loc[r+3+i,'Overall Transition Count':]=table[i]
                        r+=(3+i)
            except:
                print("error in writing the mod transition count tables to file")
                exit()
                
            try:
                longsublength = {1:0,-1:0,2:0,-2:0,3:0,-3:0,4:0,-4:0}
                # Counting longest Subsequence Length
                i=0
                while i <= c-1:
                    oct1 = df["Octant"][i]
                    cnt = 0
                    j = i
                    while j <= c-1:
                        try:
                            oct2 = df["Octant"][j]
                            if(oct2 == oct1):
                                cnt +=1 
                            else:
                                break
                            j += 1
                        except:
                            print("error in counting longest subsequence")
                    longsublength[df["Octant"][i]] = max(longsublength[df["Octant"][i]], cnt)
                    i = j 
                # Counting number of longest subsequences for each octant
                cntlongsub = {1:0,-1:0,2:0,-2:0,3:0,-3:0,4:0,-4:0}
                #  for each octant's longest subsequence Creating list of T ranges
                rangeLS = {1:[],-1:[],2:[],-2:[],3:[],-3:[],4:[],-4:[]}
                i=0        
                while i <= c-1:
                    cnt = 0
                    try:
                        oct1 = df["Octant"][i]
                        j = i
                        while j <= c-1:
                            oct2= df["Octant"][j]
                            if(oct2 == oct1):
                                cnt +=1 
                            else:
                                break
                            j += 1
                    except:
                        print("error in counting longest subsequence")
                    
                    if(cnt == longsublength[oct1]):
                        cntlongsub[oct1] += 1
                        try:                                                #creating list of T range
                            lst=[]
                            lst.append(df.loc[i,'T'])
                            lst.append(df.loc[j-1,'T'])
                            rangeLS[oct1].append(lst)
                        except:
                            print("Invalid value")            
                    i = j
            except:
                print("error while calculating longest subsequence")
                
            #writing the output of longest subsequence to file

            oct={0:'+1',1:'-1',2:'+2',3:'-2',4:'+3',5:'-3',6:'+4',7:'-4'}
            index={0:1,1:-1,2:2,3:-2,4:3,5:-3,6:4,7:-4}
            ind=2
            df['            ']=''        #12 space column
            df['Longest Subsequence Length']=''
            df['             ']=''       #13 space column
            df['              ']=''      #14 space column
            df['               ']=''     #15 space column
            df['Longest Subsequence Length with Range']=''
            df['                ']=''    #16 space column
            df['                 ']=''    #17 space column
            df.loc[1,'Longest Subsequence Length']="Octant###"
            df.loc[1,'             ']="Longest Subsequence Length"
            df.loc[1,'              ']="Count"
            df.loc[1,'Longest Subsequence Length with Range']="Octant###"
            df.loc[1,'                ']="Longest Subsequence Length"
            df.loc[1,'                 ']="Count"
            for i in range(2,10):
                df.loc[i,'Longest Subsequence Length']=oct[i-2]
                df.loc[i,'             ']=longsublength[index[i-2]]
                df.loc[i,'              ']=cntlongsub[index[i-2]]
                df.loc[ind,'Longest Subsequence Length with Range']=oct[i-2]
                df.loc[ind,'                ']=longsublength[index[i-2]]
                df.loc[ind,'                 ']=cntlongsub[index[i-2]]
                ind+=1
                df.loc[ind,'Longest Subsequence Length with Range']='T'
                df.loc[ind,'                ']='From'
                df.loc[ind,'                 ']='To'
                ind+=1
                for time in rangeLS[index[i-2]]:
                    lst=time
                    df.loc[ind, "                "]  = lst[0]
                    df.loc[ind, "                 "] = lst[1]
                    ind += 1
            os.chdir(r'C:\Users\PRIYA RAJ\OneDrive\Documents\GitHub\2001EE46_2022\proj2')
            df.to_excel('output/temp.xlsx', index = False)       #finally writing the output to the xlsx file

            fill_cell= op.styles.PatternFill(patternType='solid', fgColor='FFFF00') # to fill yellow colour
            thin = Side(border_style="thin", color="000000") # for border part     
            try: # opening the input excel sheet
                os.chdir(r'C:\Users\PRIYA RAJ\OneDrive\Documents\GitHub\2001EE46_2022\proj2\output') # making the parent directory as input directory
                inp=op.load_workbook("temp.xlsx")
                ipsheet=inp.active
            except Exception as e: # handling exception and exitting the program
                print(f"Can't open input file {file}")
                print(e)
                return

            try: # opening the input excel sheet
                try:
                    os.mkdir(outputpath)
                except:
                    pass
                os.chdir(outputpath) # setting the output directory
                output=op.Workbook()
                opsheet=output.active
            except Exception as e: # handling exception and exitting the program
                print(f"Can't open output file for input file {file}")
                print(e)
                return
            # calculate total number of rows and 
            # columns in source excel file
            mr = ipsheet.max_row
            mc = ipsheet.max_column

            # copying the cell values from source 
            # excel file to destination excel file
            for i in range (1, mr + 1):
                for j in range (1, mc + 1):
                    # reading cell value from source excel file
                    c = ipsheet.cell(row = i, column = j)

                    # writing the read value to destination excel file
                    opsheet.cell(row = i, column = j).value = c.value
            #border and color for octant count and range
            for row in range(1,4+modcount):
                for col in range(14,33):
                    opsheet.cell(row=row,column=col).border = Border(top=thin, left=thin, right=thin, bottom=thin) # giving border
                for col in range(23,31):                                        #filling color
                    if(opsheet.cell(row=row,column=col).value==1):
                        opsheet.cell(row=row,column=col).fill=fill_cell
            
            #border for table below rank and count table
            for row in range(8+modcount,modcount+17):
                for col in range(29,32):
                    opsheet.cell(row=row,column=col).border = Border(top=thin, left=thin, right=thin, bottom=thin) # giving border
                    
            #border for transition count tables
            r=3
            for i in range(modcount+2):
                for row in range(r,r+9):
                    for col in range(35,44):
                        opsheet.cell(row=row,column=col).border = Border(top=thin, left=thin, right=thin, bottom=thin) # giving border
                col=36
                for row in range(r+1,r+9):
                    opsheet.cell(row=row,column=col).fill=fill_cell
                    col+=1                   
                r+=13
            
            #border for longest subsequence table
            for row in range(3,12):
                for col in range(45,48):
                    opsheet.cell(row=row,column=col).border = Border(top=thin, left=thin, right=thin, bottom=thin) # giving border  
            
            #border for longest subsequence with ranges tables
            for row in range(3,2+ind):
                for col in range(49,52):
                    opsheet.cell(row=row,column=col).border = Border(top=thin, left=thin, right=thin, bottom=thin) # giving border                       
            os.chdir(r'C:\Users\PRIYA RAJ\OneDrive\Documents\GitHub\2001EE46_2022\output')
            os.remove("temp.xlsx")
            os.chdir(outputpath)
            output.save(file_name)
            print("Successful operation")
        global temp;
        temp=1
        
    except:
        print("error in main function")
        st.success("computation unsuccessful")
        exit()
    
#tile section
st.set_page_config(page_title="Multiple Files Computation",page_icon=":computer:",layout="wide")
st.header("Welcome to multiple files page. :v:")
st.write("If you want to convert a multiple input files into output files, then you are on the correct page. :heavy_check_mark:")

#--use local CSS
def local_css(file_name):
    with open(file_name) as f:
        st.markdown(f"<style>{f.read()}</style>",unsafe_allow_html=True)

local_css("style/style2.css")
#--input from user----
with st.container():
    st.write("---")
    st.subheader("Input path to files")
    input_path=st.text_input(r"Enter the path of the input files",placeholder=r"Please enter the path, for example:- 'C:\\Users\\PRIYA RAJ\\OneDrive\\Documents\\GitHub\\2001EE46_2022\\proj2\\output' (without inverted commas) ")

with st.container():
    st.write("---")
    st.subheader("Taking Mod value")
    mod=st.text_input("Enter the mod value:",placeholder="Please enter an integer mod value, for example:- 5000")
    try:
        mod=int(mod)
    except Exception as e:
        st.write("Please enter an integer value for mod")
        print(e)
        
    st.write("###")
    work=st.button("Compute",help="Click here for computing the file",type="primary")
    
    if work:
        try:
            os.mkdir(r'C:\Users\PRIYA RAJ\OneDrive\Documents\GitHub\2001EE46_2022\proj2\output')
        except:
            pass
        try:
            x=os.listdir(input_path)
            octant_analysis(input_path,mod)
        except:
            if input_path!="":
                st.write("Please provide a correct path!!!")
        if temp==1:
            st.success("computation of all files successful. Output files are saved in output folder on given path")