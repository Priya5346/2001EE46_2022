import os
import openpyxl
from platform import python_version
from datetime import datetime
start_time = datetime.now()

ver = python_version()

os.system("cls")

if ver == "3.8.10":
    print("Correct Version Installed")
else:
    print("Please install 3.8.10. Instruction are present in the GitHub Repo/Webmail. Url: https://pastebin.com/nvibxmjw")

inputAttendance = "input_attendance.csv"
inputRegisteredFile = "input_registered_students.csv"

roll_to_name = {}
roll_attendance = {}
dates = []

def consolidate_attendance_func():
    try:
        outputFileName = "output/attendance_report_consolidated.xlsx"
        outputFile = openpyxl.Workbook()
        outputSheet = outputFile.active

        outputSheet.cell(row=1, column=1).value = "Roll"
        outputSheet.cell(row=1, column=2).value = "Name"

        last = -1
        for i, date in enumerate(dates):
            outputSheet.cell(row=1, column=3+i).value = date
            last = i+3

        last += 1
        list = ["Actual Lecture Taken", "Total Real", "% Attendance"]

        for i, title in enumerate(list):
            outputSheet.cell(row=1, column=last+i).value = title

        for i, rollNum in enumerate(roll_to_name.keys()):
            outputSheet.cell(row=i+2, column=1).value = rollNum
            outputSheet.cell(row=i+2, column=2).value = roll_to_name[rollNum]

            present = 0
            for j, date in enumerate(dates):
                if date not in roll_attendance[rollNum]:
                    outputSheet.cell(row=i+2, column=j+3).value = "A"
                else:
                    list = roll_attendance[rollNum][date]
                    total = list[0]+list[1]+list[2]
                    if total == 0:
                        outputSheet.cell(row=i+2, column=j+3).value = "A"
                    else:
                        outputSheet.cell(row=i+2, column=j+3).value = "P"
                        present += 1

            outputSheet.cell(row=i+2, column=last).value = len(dates)
            outputSheet.cell(row=i+2, column=last+1).value = present
            percentage_attendance = (100*present)/len(dates)
            percentage_attendance = round(percentage_attendance, 2)
            outputSheet.cell(row=i+2, column=last+2).value = percentage_attendance

        outputFile.save(outputFileName)
    except:
        print("Folder named output doesn't exist")
        exit()

def roll_attendance_func():
    title = ["Date", "Roll", "Name", "Total Attendance Count",
             "Real", "Duplicate", "Invalid", "Absent"]

    for rollNum in roll_to_name.keys():
        try:
            outputFileName = "output/" + rollNum + ".xlsx"
            outputFile = openpyxl.Workbook()
            outputSheet = outputFile.active

            for i, word in enumerate(title):
                outputSheet.cell(row=1, column=i+1).value = word
            outputSheet.cell(row=2, column=2).value = rollNum
            outputSheet.cell(row=2, column=3).value = roll_to_name[rollNum]

            attendance = roll_attendance[rollNum]  # map of date -> array

            for i, date in enumerate(attendance.keys()):
                outputSheet.cell(row=3+i, column=1).value = date
                list = attendance[date]
                total = list[0]+list[1]+list[2]

                outputSheet.cell(row=3+i, column=4).value = total
                outputSheet.cell(row=3+i, column=5).value = list[0]
                outputSheet.cell(row=3+i, column=6).value = list[1]
                outputSheet.cell(row=3+i, column=7).value = list[2]

                if total == 0:
                    outputSheet.cell(row=3+i, column=8).value = 1
                else:
                    outputSheet.cell(row=3+i, column=8).value = 0

            outputFile.save(outputFileName)
        except:
            print("The folder output doesn't exist")
            exit()


def valid_day(date):
    day_rank = datetime.strptime(date, '%d-%m-%Y').weekday()

    if day_rank == 0 or day_rank == 3:
        return True
    return False


def valid_time(time):
    
    hrs, minute = time.split(":")
    hrs = int(hrs)
    minute = int(minute)
    if hrs == 15 and minute == 0:
        return True
    if hrs == 14:
        return True
    return False


def attendance_count_func():
    try:
        
        f = open(inputAttendance, "r")

        f.readline()

        all_lines = f.readlines()

        for rolls in roll_to_name.keys():
            roll_attendance[rolls] = {}

        for line in all_lines:
            line = line.strip()
            timestamp, naming = line.split(",")
            date, time = timestamp.split(" ")
            rollNumber = naming[:8]

            if valid_day(date):
                if date not in dates:
                    dates.append(date)

                if date not in roll_attendance[rollNumber]:
                    roll_attendance[rollNumber][date] = [0, 0, 0]

                if valid_time(time):
                    if (roll_attendance[rollNumber][date][0] == 0):
                        roll_attendance[rollNumber][date][0] = 1
                    else:
                        roll_attendance[rollNumber][date][1] += 1

                else:
                    roll_attendance[rollNumber][date][2] += 1
    except FileNotFoundError:
        print('File not found')
        exit()


def map_roll_to_num_func():
    try:
        
        f = open(inputRegisteredFile, "r")
        
        f.readline()
        
        all_lines = f.readlines()

        for line in all_lines:
            line = line.strip()
            list = line.split(",")

            rollNo = list[0].strip()
            name = list[1].strip()
            roll_to_name[rollNo] = name

        f.close()
    except FileNotFoundError:
        print("File not found")
        exit()


def attendance_report():
    try:
        
        map_roll_to_num_func()
    except NameError:
        print('Either the function is not created or the name is not correct.')
        exit()
    try:
        
        attendance_count_func()
    except NameError:
        print('Either the function is not created or the name is not correct.')
        exit()
    try:
        
        roll_attendance_func()
    except NameError:
        print('Either the function is not created or the name is not correct.')
        exit()
    try:
        
        consolidate_attendance_func()
    except NameError:
        print('Either the function is not created or the name is not correct.')
        exit()

try:
    attendance_report()
except NameError:
    print('Either the function is not created or the name is not correct.')
    exit()


end_time = datetime.now()
print('Duration of Program Execution: {}'.format(end_time - start_time))
